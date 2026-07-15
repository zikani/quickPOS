import os
from datetime import datetime

def export_pdf_summary_file(filepath: str, summary_data: dict, cashiers_data: list, top_products_data: list) -> bool:
    """Export standard high-fidelity financial summary statements PDF using reportlab."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=15
        )
        
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1e293b'),
            spaceBefore=15,
            spaceAfter=8
        )
        
        body_style = styles['BodyText']
        
        # Title & Subtitle
        story.append(Paragraph("QuickPOS - Executive Financial Report", title_style))
        story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
        story.append(Spacer(1, 15))
        
        # Summary Grid
        story.append(Paragraph("1. Core Business Metrics", section_style))
        metrics_data = [
            ["Business Metric", "Reported Value"],
            ["Total Sales Volume", f"{summary_data['transaction_count']} orders"],
            ["Gross Revenue", f"${summary_data['total_revenue']:.2f}"],
            ["Total Tax Collected", f"${summary_data['total_tax']:.2f}"],
            ["Total Discounts Issued", f"${summary_data['total_discount']:.2f}"],
            ["Net Calculated Profit", f"${summary_data['net_profit']:.2f}"],
            ["Profit Margin %", f"{summary_data['profit_margin_pct']:.2f}%"]
        ]
        
        t_metrics = Table(metrics_data, colWidths=[240, 200])
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8fafc')),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        story.append(t_metrics)
        story.append(Spacer(1, 15))
        
        # Cashier Performance
        story.append(Paragraph("2. Cashier Operations Breakdown", section_style))
        cashier_rows = [["Operator Name", "Ticket count", "Revenue Value"]]
        for c in cashiers_data:
            cashier_rows.append([c['name'], f"{c['sales_count']} sales", f"${c['revenue_total']:.2f}"])
            
        t_cashier = Table(cashier_rows, colWidths=[180, 130, 130])
        t_cashier.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#475569')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        story.append(t_cashier)
        story.append(Spacer(1, 15))
        
        # Bestselling Catalog Items
        story.append(Paragraph("3. Best-Selling Catalog Items", section_style))
        prod_rows = [["Product Name", "Units Sold", "Total sales"]]
        for p in top_products_data:
            prod_rows.append([p['name'], f"{p['total_qty']} units", f"${p['total_sales']:.2f}"])
            
        t_prod = Table(prod_rows, colWidths=[180, 130, 130])
        t_prod.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#16a34a')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        story.append(t_prod)
        
        doc.build(story)
        print(f"[+] PDF Summary written to {filepath}")
        return True
    except Exception as e:
        print(f"[!] PDF generation failed: {e}. Writing formatted plain-text fallback.")
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("="*50 + "\n")
                f.write("      QUICKPOS EXECUTIVE FINANCIAL SUMMARY\n")
                f.write(f"      Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("="*50 + "\n\n")
                
                f.write("1. CORE METRICS:\n")
                f.write(f" - Transactions: {summary_data['transaction_count']}\n")
                f.write(f" - Gross Revenue: ${summary_data['total_revenue']:.2f}\n")
                f.write(f" - Total Tax: ${summary_data['total_tax']:.2f}\n")
                f.write(f" - Total Discounts: ${summary_data['total_discount']:.2f}\n")
                f.write(f" - Net Profit: ${summary_data['net_profit']:.2f}\n")
                f.write(f" - Margin: {summary_data['profit_margin_pct']:.2f}%\n\n")
                
                f.write("2. CASHIERS BREAKDOWN:\n")
                for c in cashiers_data:
                    f.write(f" - {c['name']}: {c['sales_count']} sales, ${c['revenue_total']:.2f}\n")
                f.write("\n")
                
                f.write("3. TOP SELLING PRODUCTS:\n")
                for p in top_products_data:
                    f.write(f" - {p['name']}: {p['total_qty']} units, ${p['total_sales']:.2f}\n")
                f.write("\n" + "="*50 + "\n")
            return True
        except Exception as inner_e:
            print(f"[!] Fallback output write failed: {inner_e}")
            return False

def export_excel_summary_file(filepath: str, summary_data: dict, cashiers_data: list, top_products_data: list) -> bool:
    """Export multi-tab formatted workbook sheet using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # ---------------- SHEET 1: Core Summary ----------------
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        # Title block
        ws1["A1"] = "QuickPOS Core Financials"
        ws1["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F172A")
        ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws1["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")
        
        # Headers
        ws1.cell(row=4, column=1, value="Financial Metric").font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=4, column=1).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        ws1.cell(row=4, column=2, value="Value").font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=4, column=2).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        
        metrics = [
            ["Total Sales Volume", summary_data['transaction_count']],
            ["Gross Revenue", summary_data['total_revenue']],
            ["Total Tax", summary_data['total_tax']],
            ["Total Discount Issued", summary_data['total_discount']],
            ["Net Profit", summary_data['net_profit']],
            ["Margin %", f"{summary_data['profit_margin_pct']:.2f}%"]
        ]
        
        for idx, (m_title, m_val) in enumerate(metrics, 5):
            ws1.cell(row=idx, column=1, value=m_title)
            ws1.cell(row=idx, column=2, value=m_val)
            
        # Adjust widths
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 15
        
        # ---------------- SHEET 2: Operational Breakdowns ----------------
        ws2 = wb.create_sheet(title="Operator & Product Audits")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "Cashier Performance Metrics"
        ws2["A1"].font = Font(size=12, bold=True, color="1E293B")
        
        c_headers = ["Operator Name", "Ticket count", "Revenue Value"]
        for col_idx, h in enumerate(c_headers, 1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            
        curr_row = 4
        for c in cashiers_data:
            ws2.cell(row=curr_row, column=1, value=c['name'])
            ws2.cell(row=curr_row, column=2, value=c['sales_count'])
            ws2.cell(row=curr_row, column=3, value=c['revenue_total'])
            curr_row += 1
            
        # Add products block
        curr_row += 2
        ws2.cell(row=curr_row, column=1, value="Bestselling Catalog Products").font = Font(size=12, bold=True, color="1E293B")
        
        p_headers = ["Product Name", "Units Sold", "Total sales Value"]
        for col_idx, h in enumerate(p_headers, 1):
            cell = ws2.cell(row=curr_row+2, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
            
        start_row = curr_row + 3
        for p in top_products_data:
            ws2.cell(row=start_row, column=1, value=p['name'])
            ws2.cell(row=start_row, column=2, value=p['total_qty'])
            ws2.cell(row=start_row, column=3, value=p['total_sales'])
            start_row += 1
            
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 20
        
        wb.save(filepath)
        print(f"[+] Excel workbook written to {filepath}")
        return True
    except Exception as e:
        print(f"[!] Excel workbook creation failed: {e}. Writing formatted CSV fallback.")
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("Section,Metric Name,Value\n")
                f.write(f"Summary,Total Sales Volume,{summary_data['transaction_count']}\n")
                f.write(f"Summary,Gross Revenue,{summary_data['total_revenue']:.2f}\n")
                f.write(f"Summary,Net Profit,{summary_data['net_profit']:.2f}\n")
                for c in cashiers_data:
                    f.write(f"Cashier Performance,{c['name']},{c['revenue_total']:.2f}\n")
                for p in top_products_data:
                    f.write(f"Bestseller,{p['name']},{p['total_sales']:.2f}\n")
            return True
        except Exception as inner_e:
            print(f"[!] Fallback CSV write failed: {inner_e}")
            return False
